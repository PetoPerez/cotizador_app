import io
import uuid
from dataclasses import dataclass
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, selectinload
from app.database import get_db
from app import schemas
from app.security import get_current_user, require_admin
from app import models
from app.services.storage_service import upload_image, delete_image, key_from_url
from app.services.precio_audit import registrar_cambio_precio, ref_producto

router = APIRouter(prefix="/productos", tags=["productos"])

# Encabezados base del Excel (normalizados a lowercase)
_COL_BASE = {
    "marca":       "marca",
    "equipo":      "equipo",
    "modelo":      "modelo",
    "descripcion": "descripcion",
    "descripción": "descripcion",
}


def _empresas_para_import(db: Session) -> list[models.Empresa]:
    """Empresas que aceptan productos importables (todas excepto servicios_lavanderia)."""
    return (db.query(models.Empresa)
              .filter(models.Empresa.codigo != 'servicios_lavanderia',
                      models.Empresa.activa == True)
              .order_by(models.Empresa.nombre)
              .all())


@router.get("/", response_model=list[schemas.ProductoOut])
def listar(
    q: str = None,
    empresa: str = None,  # filtro opcional: código de empresa (clm, supliese_gamesail, etc.)
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    query = (db.query(models.Producto)
               .options(selectinload(models.Producto.imagenes),
                        selectinload(models.Producto.empresas))
               .filter(models.Producto.activo == True))
    if q:
        query = query.filter(
            models.Producto.modelo.ilike(f"%{q}%") |
            models.Producto.marca.ilike(f"%{q}%") |
            models.Producto.equipo.ilike(f"%{q}%")
        )
    if empresa:
        empresa_obj = db.query(models.Empresa).filter(models.Empresa.codigo == empresa).first()
        if empresa_obj:
            query = (query.join(models.ProductoEmpresa,
                                models.ProductoEmpresa.producto_id == models.Producto.id)
                          .filter(models.ProductoEmpresa.empresa_id == empresa_obj.id,
                                  models.ProductoEmpresa.activo == True))
    return query.order_by(models.Producto.marca, models.Producto.equipo).all()


@router.get("/plantilla-importar")
def descargar_plantilla(db: Session = Depends(get_db), _=Depends(require_admin)):
    """Descarga un Excel vacío con los encabezados correctos para importar productos."""
    empresas = _empresas_para_import(db)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos"

    # Encabezados base + precio_general + una columna por empresa
    headers = ["marca", "equipo", "modelo", "descripcion", "precio_general"]
    for emp in empresas:
        headers.append(f"precio_{emp.acronimo.lower()}")

    # Estilo encabezado
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="26326E", end_color="26326E", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")

    for col_idx, name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        ws.column_dimensions[cell.column_letter].width = 18

    # Fila de ejemplo: producto normal con precio por empresa
    ejemplo = ["GIRBAU", "Lavadora industrial", "HS-6028", "Capacidad 28kg, motor inverter", ""]
    for emp in empresas:
        ejemplo.append(75000)
    for col_idx, val in enumerate(ejemplo, start=1):
        ws.cell(row=2, column=col_idx, value=val)

    # Instrucciones
    nota = ("Los precios se capturan en USD (dólares). · Llena solo los precios que "
            "apliquen. · precio_general: el producto queda disponible para TODAS las "
            "empresas a ese precio (si dejas la marca vacía se pone 'General'). · Los "
            "servicios se importan desde su propia plantilla (pantalla de Servicios).")
    ws.cell(row=3, column=1, value=nota)
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(headers))
    ws.cell(row=3, column=1).font = Font(italic=True, color="808080")
    ws.cell(row=3, column=1).alignment = Alignment(horizontal="left")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="plantilla_productos.xlsx"'},
    )


def _parsear_precio_import(raw):
    if raw is None or raw == "":
        return None
    try:
        v = float(str(raw).replace(",", "").replace("$", "").strip())
        return v if v > 0 else None
    except (ValueError, TypeError):
        return None


@dataclass
class _ProductoExistente:
    """Vista mínima de un producto ya guardado, para emparejar en la importación.
    precios: {codigo_empresa: (precio_lista|None, activo)}."""
    id: object
    descripcion: object
    precios: dict


def _extension_valida(filename) -> bool:
    """Acepta .xlsx/.xls sin importar mayúsculas ('LISTA.XLSX' es válido)."""
    return bool(filename) and str(filename).lower().strip().endswith((".xlsx", ".xls"))


def _norm_clave(s) -> str:
    """Normaliza marca/equipo/modelo para emparejar: minúsculas y espacios
    colapsados. Así 'GIRBAU', 'girbau' y ' Girbau ' son el MISMO producto y la
    importación ACTUALIZA en vez de crear un duplicado."""
    return " ".join(str(s or "").split()).lower()


def _analizar_importacion(rows, empresas, catalogo):
    """Clasifica las filas del Excel en nuevos / a actualizar / sin cambios / errores.

    Función PURA (sin base de datos) para poder probarla en aislamiento; ver
    tests/test_import_productos.py.

      rows     : filas del Excel (incluye la de encabezados), como devuelve
                 openpyxl iter_rows(values_only=True).
      empresas : objetos con .codigo y .acronimo (empresas importables).
      catalogo : {(marca,equipo,modelo) normalizados con _norm_clave: _ProductoExistente}.

    Devuelve dict con `error` (str|None; problema de formato: si no es None nada
    debe aplicarse) y, en éxito, nuevos / actualizar / sin_cambios / errores /
    sin_marca / columnas_ignoradas.
    """
    vacio = {"error": None, "nuevos": [], "actualizar": [], "sin_cambios": 0,
             "errores": [], "sin_marca": 0, "columnas_ignoradas": []}

    if not rows:
        return {**vacio, "error": "El archivo está vacío"}

    empresas_por_codigo = {e.codigo: e for e in empresas}
    precio_col_to_empresa = {f"precio_{e.acronimo.lower()}": e for e in empresas}

    # ── Validación de formato (encabezados). Si está mal, NO se carga nada. ──
    headers = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
    col_idx = {}
    precio_col_idx = {}
    idx_general = None
    cols_desconocidas = []
    for i, h in enumerate(headers):
        if not h:
            continue
        if h in _COL_BASE:
            col_idx[_COL_BASE[h]] = i
        elif h == "precio_general":
            idx_general = i
        elif h in precio_col_to_empresa:
            precio_col_idx[precio_col_to_empresa[h].codigo] = i
        elif h in ("precio_servicios", "precio_sdl"):
            # Los servicios ahora tienen su propia plantilla de importación.
            cols_desconocidas.append(h + " (los servicios se importan por separado)")
        else:
            cols_desconocidas.append(h)

    missing = {"equipo", "modelo"} - col_idx.keys()
    if missing:
        return {**vacio, "columnas_ignoradas": cols_desconocidas,
                "error": f"Formato incorrecto: faltan columnas requeridas "
                         f"({', '.join(sorted(missing))}). Descarga la plantilla y respeta "
                         f"los encabezados. Encabezados detectados: "
                         f"{', '.join(h for h in headers if h) or '(ninguno)'}"}
    if not precio_col_idx and idx_general is None:
        return {**vacio, "columnas_ignoradas": cols_desconocidas,
                "error": "Formato incorrecto: incluye al menos una columna de precio "
                         "(precio_general, precio_clm, precio_gs, precio_sup o precio_gir)."}

    def cell(row, idx):
        return row[idx] if idx is not None and idx < len(row) else None

    nuevos = []       # {marca,equipo,modelo,desc,precios,sin_marca_fila,referencia}
    actualizar = []   # {existente_id,desc,precios,cambia_desc,cambios[str],referencia}
    sin_cambios = 0
    errores = []
    sin_marca = 0

    for n, row in enumerate(rows[1:], start=2):
        marca_raw = str(cell(row, col_idx.get("marca")) or "").strip()
        sin_marca_fila = (marca_raw == "")
        marca  = marca_raw or "General"
        equipo = str(cell(row, col_idx.get("equipo")) or "").strip()
        modelo = str(cell(row, col_idx.get("modelo")) or "").strip()
        desc   = str(cell(row, col_idx.get("descripcion")) or "").strip() or None

        if not equipo and not modelo:
            continue  # fila en blanco
        if not equipo or not modelo:
            errores.append(f"Fila {n}: falta equipo o modelo")
            continue

        precios_fila = {}
        if idx_general is not None:
            pg = _parsear_precio_import(cell(row, idx_general))
            if pg is not None:
                for e in empresas:
                    precios_fila[e.codigo] = pg
        for codigo, idx in precio_col_idx.items():
            p = _parsear_precio_import(cell(row, idx))
            if p is not None:
                precios_fila[codigo] = p

        if not precios_fila:
            errores.append(f"Fila {n} ({modelo}): sin ningún precio válido")
            continue

        ref = f"{marca} / {equipo} / {modelo}"
        clave = (_norm_clave(marca), _norm_clave(equipo), _norm_clave(modelo))
        existente = catalogo.get(clave)

        if existente:
            cambios = []
            cambia_desc = desc is not None and (desc or None) != (existente.descripcion or None)
            if cambia_desc:
                cambios.append("descripción")
            for codigo, precio in precios_fila.items():
                emp = empresas_por_codigo[codigo]
                antes_precio, antes_activo = existente.precios.get(codigo, (None, False))
                antes = float(antes_precio) if (antes_precio is not None and antes_activo) else None
                if antes is None or round(antes, 2) != round(precio, 2):
                    antes_txt = f"{antes:,.2f}" if antes is not None else "—"
                    cambios.append(f"precio {emp.acronimo}: {antes_txt} → {precio:,.2f}")
            if cambios:
                actualizar.append({"existente_id": existente.id, "desc": desc,
                                   "precios": precios_fila, "cambia_desc": cambia_desc,
                                   "cambios": cambios, "referencia": ref})
            else:
                sin_cambios += 1
        else:
            if sin_marca_fila:
                sin_marca += 1
            nuevos.append({"marca": marca, "equipo": equipo, "modelo": modelo, "desc": desc,
                           "precios": precios_fila, "sin_marca_fila": sin_marca_fila,
                           "referencia": ref})

    return {"error": None, "nuevos": nuevos, "actualizar": actualizar,
            "sin_cambios": sin_cambios, "errores": errores, "sin_marca": sin_marca,
            "columnas_ignoradas": cols_desconocidas}


@router.post("/importar")
def importar_excel(
    file: UploadFile = File(...),
    confirmar: bool = False,       # False = solo vista previa (no escribe); True = aplica
    marca_general: bool = False,   # confirma guardar productos sin marca como 'General'
    db: Session = Depends(get_db),
    current_user: models.Usuario = Depends(require_admin),
):
    """Importa productos desde Excel en dos fases.

    Fase 1 (confirmar=False): analiza el archivo SIN escribir y devuelve una vista
    previa (nuevos, a actualizar con el detalle de cambios, sin cambios, errores).
    Fase 2 (confirmar=True): aplica solo lo nuevo y lo que realmente cambió.
    """
    if not _extension_valida(file.filename):
        raise HTTPException(status_code=400, detail="Solo se aceptan archivos .xlsx o .xls")
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file.file.read()), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="No se pudo leer el archivo Excel")

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    empresas = _empresas_para_import(db)  # CLM, GS, SUP, GIR (sin servicios)
    empresas_por_codigo = {e.codigo: e for e in empresas}
    emp_por_id = {str(e.id): e for e in empresas}

    # Índice de productos existentes en UNA sola consulta. El emparejamiento usa la
    # clave normalizada (minúsculas/espacios colapsados), de modo que si el Excel
    # trae otra caja o espaciado que la BD el producto se ACTUALIZA en vez de
    # duplicarse. (Antes el match era exacto y sensible a mayúsculas.)
    existentes_db = (db.query(models.Producto)
                       .options(selectinload(models.Producto.empresas)).all())
    catalogo = {}
    orm_por_id = {}
    for p in existentes_db:
        clave = (_norm_clave(p.marca), _norm_clave(p.equipo), _norm_clave(p.modelo))
        precios = {}
        for pe in p.empresas:
            emp = emp_por_id.get(str(pe.empresa_id))
            if emp:
                precios[emp.codigo] = (pe.precio_lista, pe.activo)
        catalogo[clave] = _ProductoExistente(id=str(p.id), descripcion=p.descripcion, precios=precios)
        orm_por_id[str(p.id)] = p

    # ── Clasificación (función pura, cubierta por tests/test_import_productos.py) ──
    resultado = _analizar_importacion(rows, empresas, catalogo)
    if resultado["error"]:
        raise HTTPException(status_code=400, detail=resultado["error"])

    nuevos = resultado["nuevos"]
    actualizar = resultado["actualizar"]
    sin_cambios = resultado["sin_cambios"]
    errores = resultado["errores"]
    sin_marca = resultado["sin_marca"]

    resumen = {"nuevos": len(nuevos), "actualizar": len(actualizar),
               "sin_cambios": sin_cambios, "errores": len(errores), "sin_marca": sin_marca}
    preview = {
        "resumen": resumen,
        "nuevos": [f"Fila nueva: {x['referencia']}" + (" (sin marca → 'General')" if x["sin_marca_fila"] else "")
                   for x in nuevos[:100]],
        "actualizar": [{"item": x["referencia"], "cambios": x["cambios"]} for x in actualizar[:100]],
        "errores": errores[:100],
        "columnas_ignoradas": resultado["columnas_ignoradas"],
    }

    # ── Fase 1: vista previa ──
    if not confirmar:
        return {**preview, "confirmado": False}

    # Guarda de productos sin marca: exige confirmación explícita adicional.
    if sin_marca > 0 and not marca_general:
        return {**preview, "confirmado": False, "requiere_marca_general": True}

    # ── Fase 2: aplicar (solo nuevos + cambios reales) ──
    def upsert_precio(producto, empresa, precio):
        existentes = {str(pe.empresa_id): pe for pe in producto.empresas}
        eid = str(empresa.id)
        if eid in existentes:
            anterior = existentes[eid].precio_lista
            existentes[eid].precio_lista = precio
            existentes[eid].activo = True
        else:
            anterior = None
            db.add(models.ProductoEmpresa(producto_id=producto.id, empresa_id=empresa.id,
                                          precio_lista=precio, activo=True))
        registrar_cambio_precio(
            db, tipo="producto", referencia=ref_producto(producto, empresa),
            precio_nuevo=precio, precio_anterior=anterior,
            producto_id=producto.id, empresa_id=empresa.id,
            usuario=current_user, origen="importacion")

    try:
        for x in nuevos:
            p = models.Producto(marca=x["marca"], equipo=x["equipo"], modelo=x["modelo"], descripcion=x["desc"])
            db.add(p)
            db.flush()
            for codigo, precio in x["precios"].items():
                upsert_precio(p, empresas_por_codigo[codigo], precio)
        for x in actualizar:
            p = orm_por_id[str(x["existente_id"])]
            if x["cambia_desc"]:
                p.descripcion = x["desc"]
            db.flush()
            for codigo, precio in x["precios"].items():
                upsert_precio(p, empresas_por_codigo[codigo], precio)
        db.commit()
    except Exception as ex:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"Error aplicando la importación: {ex}")

    return {**preview, "confirmado": True}


@router.post("/", response_model=schemas.ProductoOut)
def crear(data: schemas.ProductoCreate, db: Session = Depends(get_db),
          current_user: models.Usuario = Depends(require_admin)):
    if not data.empresas:
        raise HTTPException(status_code=400, detail="Debes asignar el producto a al menos una empresa")

    producto = models.Producto(
        marca=data.marca, equipo=data.equipo, modelo=data.modelo,
        descripcion=data.descripcion,
    )
    db.add(producto)
    db.flush()  # asigna id

    empresas_map = {str(e.id): e for e in db.query(models.Empresa)
                    .filter(models.Empresa.id.in_([pe.empresa_id for pe in data.empresas])).all()}
    for pe in data.empresas:
        db.add(models.ProductoEmpresa(
            producto_id=producto.id,
            empresa_id=pe.empresa_id,
            precio_lista=pe.precio_lista,
            activo=pe.activo,
        ))
        emp = empresas_map.get(str(pe.empresa_id))
        if emp:
            registrar_cambio_precio(
                db, tipo="producto", referencia=ref_producto(producto, emp),
                precio_nuevo=pe.precio_lista, producto_id=producto.id,
                empresa_id=pe.empresa_id, usuario=current_user, origen="manual",
            )
    db.commit()
    db.refresh(producto)
    return producto


@router.put("/{id}", response_model=schemas.ProductoOut)
def actualizar(id: str, data: schemas.ProductoUpdate, db: Session = Depends(get_db),
               current_user: models.Usuario = Depends(require_admin)):
    producto = db.query(models.Producto).filter(models.Producto.id == id).first()
    if not producto:
        raise HTTPException(status_code=404, detail="Producto no encontrado")

    payload = data.model_dump(exclude_none=True)
    empresas_input = payload.pop("empresas", None)

    for field, value in payload.items():
        setattr(producto, field, value)

    if empresas_input is not None:
        # Reemplazar mapping producto_empresa con lo recibido
        existentes = {str(pe.empresa_id): pe for pe in producto.empresas}
        nuevos_ids = {str(pe["empresa_id"]) for pe in empresas_input}
        empresas_map = {str(e.id): e for e in db.query(models.Empresa)
                        .filter(models.Empresa.id.in_([pe["empresa_id"] for pe in empresas_input])).all()}

        # actualizar / insertar
        for pe in empresas_input:
            eid = str(pe["empresa_id"])
            if eid in existentes:
                anterior = existentes[eid].precio_lista
                existentes[eid].precio_lista = pe["precio_lista"]
                existentes[eid].activo = pe["activo"]
            else:
                anterior = None
                db.add(models.ProductoEmpresa(
                    producto_id=producto.id,
                    empresa_id=pe["empresa_id"],
                    precio_lista=pe["precio_lista"],
                    activo=pe["activo"],
                ))
            emp = empresas_map.get(eid)
            if emp:
                registrar_cambio_precio(
                    db, tipo="producto", referencia=ref_producto(producto, emp),
                    precio_nuevo=pe["precio_lista"], precio_anterior=anterior,
                    producto_id=producto.id, empresa_id=pe["empresa_id"],
                    usuario=current_user, origen="manual",
                )
        # eliminar los que ya no vienen
        for eid, pe in existentes.items():
            if eid not in nuevos_ids:
                db.delete(pe)

    db.commit()
    db.refresh(producto)
    return producto


@router.post("/{id}/imagen", response_model=schemas.ProductoOut)
async def subir_imagen(
    id: str,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _=Depends(require_admin),
):
    producto = db.query(models.Producto).filter(models.Producto.id == id).first()
    if not producto:
        raise HTTPException(status_code=404, detail="Producto no encontrado")
    if not (file.content_type or "").startswith("image/"):
        raise HTTPException(status_code=400, detail="Solo se aceptan archivos de imagen")

    ext = (file.filename or "img").rsplit(".", 1)[-1].lower()
    imagen_id = str(uuid.uuid4())
    key = f"productos/{id}/{imagen_id}.{ext}"
    content = await file.read()

    try:
        url = upload_image(content, key, file.content_type or "image/jpeg")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al subir imagen: {e}")

    orden = db.query(models.ProductoImagen).filter(models.ProductoImagen.producto_id == id).count()
    db.add(models.ProductoImagen(producto_id=id, url=url, orden=orden))
    if orden == 0:
        producto.imagen_url = url
    db.commit()
    db.refresh(producto)
    return producto


@router.delete("/{id}/imagen/{imagen_id}")
def eliminar_imagen(
    id: str,
    imagen_id: str,
    db: Session = Depends(get_db),
    _=Depends(require_admin),
):
    imagen = db.query(models.ProductoImagen).filter(
        models.ProductoImagen.id == imagen_id,
        models.ProductoImagen.producto_id == id,
    ).first()
    if not imagen:
        raise HTTPException(status_code=404, detail="Imagen no encontrada")

    try:
        delete_image(key_from_url(imagen.url))
    except Exception:
        pass

    db.delete(imagen)

    producto = db.query(models.Producto).filter(models.Producto.id == id).first()
    restantes = (db.query(models.ProductoImagen)
                 .filter(models.ProductoImagen.producto_id == id)
                 .order_by(models.ProductoImagen.orden).all())
    producto.imagen_url = restantes[0].url if restantes else None
    db.commit()
    return {"detail": "Imagen eliminada"}


@router.delete("/{id}")
def eliminar(id: str, db: Session = Depends(get_db), _=Depends(require_admin)):
    producto = db.query(models.Producto).filter(models.Producto.id == id).first()
    if not producto:
        raise HTTPException(status_code=404, detail="Producto no encontrado")
    producto.activo = False
    db.commit()
    return {"detail": "Producto desactivado"}
