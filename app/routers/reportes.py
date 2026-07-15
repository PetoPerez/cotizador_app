import io
from datetime import datetime, timezone

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, selectinload

from app.database import get_db
from app import schemas, models
from app.security import require_admin

router = APIRouter(prefix="/reportes", tags=["reportes"])

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="26326E", end_color="26326E", fill_type="solid")
_CENTER = Alignment(horizontal="center", vertical="center")


def _xlsx_response(wb, filename: str) -> StreamingResponse:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _estilar_encabezados(ws, ncols: int):
    for col_idx in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER
        ws.column_dimensions[get_column_letter(col_idx)].width = 20


# ── Reporte de inventario (catálogo con precios por empresa) ──
@router.get("/inventario")
def reporte_inventario(
    incluir_inactivos: bool = Query(False, description="Incluir productos desactivados"),
    empresa: str = Query(None, description="Filtrar por código de empresa (clm, supliese, ...)"),
    db: Session = Depends(get_db),
    _=Depends(require_admin),
):
    """Genera un Excel del catálogo de productos con una columna de precio por
    cada empresa. Es el 'inventario' del sistema: no maneja existencias, sino el
    catálogo cotizable y sus precios vigentes."""
    empresas = (db.query(models.Empresa)
                  .filter(models.Empresa.activa == True)
                  .order_by(models.Empresa.nombre).all())
    if empresa:
        empresas = [e for e in empresas if e.codigo == empresa] or empresas

    query = (db.query(models.Producto)
               .options(selectinload(models.Producto.empresas)))
    if not incluir_inactivos:
        query = query.filter(models.Producto.activo == True)
    productos = query.order_by(models.Producto.marca,
                               models.Producto.equipo,
                               models.Producto.modelo).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario"

    headers = ["Marca", "Equipo", "Modelo", "Descripción"]
    headers += [f"Precio {e.acronimo}" for e in empresas]
    headers += ["Estado"]
    ws.append(headers)
    _estilar_encabezados(ws, len(headers))

    for p in productos:
        precios = {str(pe.empresa_id): pe for pe in p.empresas}
        fila = [p.marca, p.equipo, p.modelo, p.descripcion or ""]
        for e in empresas:
            pe = precios.get(str(e.id))
            fila.append(float(pe.precio_lista) if pe and pe.activo else None)
        fila.append("Activo" if p.activo else "Inactivo")
        ws.append(fila)

    # Formato de moneda en las columnas de precio
    primera_precio = 5
    for col_idx in range(primera_precio, primera_precio + len(empresas)):
        for row_idx in range(2, ws.max_row + 1):
            ws.cell(row=row_idx, column=col_idx).number_format = '#,##0.00'

    ws.freeze_panes = "A2"
    fecha = datetime.now(timezone.utc).strftime("%Y%m%d")
    return _xlsx_response(wb, f"inventario_{fecha}.xlsx")


# ── Historial de cambios de precio ──
def _query_historial(db: Session, tipo, origen, desde, hasta, q):
    query = db.query(models.PrecioHistorial)
    if tipo:
        query = query.filter(models.PrecioHistorial.tipo == tipo)
    if origen:
        query = query.filter(models.PrecioHistorial.origen == origen)
    if q:
        query = query.filter(models.PrecioHistorial.referencia.ilike(f"%{q}%"))
    if desde:
        query = query.filter(models.PrecioHistorial.created_at >= desde)
    if hasta:
        query = query.filter(models.PrecioHistorial.created_at <= hasta)
    return query.order_by(models.PrecioHistorial.created_at.desc())


@router.get("/precios/historial", response_model=list[schemas.PrecioHistorialOut])
def historial_precios(
    tipo: str = Query(None, pattern="^(producto|servicio)$"),
    origen: str = Query(None),
    q: str = Query(None, description="Busca en la referencia (marca/modelo/servicio)"),
    desde: datetime = Query(None),
    hasta: datetime = Query(None),
    limit: int = Query(200, ge=1, le=1000),
    offset: int = Query(0, ge=0),
    db: Session = Depends(get_db),
    _=Depends(require_admin),
):
    """Lista los cambios de precio, del más reciente al más antiguo."""
    return (_query_historial(db, tipo, origen, desde, hasta, q)
            .offset(offset).limit(limit).all())


@router.get("/precios/historial.xlsx")
def historial_precios_excel(
    tipo: str = Query(None, pattern="^(producto|servicio)$"),
    origen: str = Query(None),
    q: str = Query(None),
    desde: datetime = Query(None),
    hasta: datetime = Query(None),
    db: Session = Depends(get_db),
    _=Depends(require_admin),
):
    registros = _query_historial(db, tipo, origen, desde, hasta, q).limit(10000).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cambios de precio"

    headers = ["Fecha", "Tipo", "Referencia", "Precio anterior",
               "Precio nuevo", "Usuario", "Origen"]
    ws.append(headers)
    _estilar_encabezados(ws, len(headers))

    for r in registros:
        ws.append([
            r.created_at.astimezone(timezone.utc).strftime("%Y-%m-%d %H:%M") if r.created_at else "",
            r.tipo,
            r.referencia,
            float(r.precio_anterior) if r.precio_anterior is not None else None,
            float(r.precio_nuevo),
            r.usuario_nombre or "—",
            r.origen,
        ])

    for col_idx in (4, 5):
        for row_idx in range(2, ws.max_row + 1):
            ws.cell(row=row_idx, column=col_idx).number_format = '#,##0.00'

    ws.freeze_panes = "A2"
    fecha = datetime.now(timezone.utc).strftime("%Y%m%d")
    return _xlsx_response(wb, f"cambios_precio_{fecha}.xlsx")
