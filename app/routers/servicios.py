import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from app.database import get_db
from app import schemas, models
from app.security import get_current_user, require_admin
from app.services.precio_audit import registrar_cambio_precio

router = APIRouter(prefix="/servicios", tags=["servicios"])

_TIPOS_VALIDOS = {"mantenimiento", "puesta_en_marcha", "otro"}
_TIPO_ALIAS = {
    "mantenimiento": "mantenimiento", "mantto": "mantenimiento",
    "puesta_en_marcha": "puesta_en_marcha", "puesta en marcha": "puesta_en_marcha",
    "puesta": "puesta_en_marcha", "puesta_marcha": "puesta_en_marcha",
    "otro": "otro", "otros": "otro",
}


def _require_sdl_o_admin(user: models.Usuario, db: Session):
    """Solo admin o vendedor asignado a Servicios de Lavandería pueden gestionar."""
    if user.rol == "admin":
        return
    if user.rol == "vendedor" and user.empresa_id:
        emp = db.query(models.Empresa).filter(models.Empresa.id == user.empresa_id).first()
        if emp and emp.codigo == "servicios_lavanderia":
            return
    raise HTTPException(status_code=403, detail="Solo admin o vendedores de Servicios de Lavandería pueden gestionar servicios")


@router.get("/", response_model=list[schemas.ServicioOut])
def listar(q: str = None, db: Session = Depends(get_db), _=Depends(get_current_user)):
    query = db.query(models.Servicio).filter(models.Servicio.activo == True)
    if q:
        query = query.filter(models.Servicio.nombre.ilike(f"%{q}%"))
    return query.order_by(models.Servicio.nombre).all()


@router.post("/", response_model=schemas.ServicioOut)
def crear(data: schemas.ServicioCreate, db: Session = Depends(get_db),
          current_user: models.Usuario = Depends(get_current_user)):
    _require_sdl_o_admin(current_user, db)
    servicio = models.Servicio(
        nombre=data.nombre,
        descripcion=data.descripcion,
        precio_unitario=data.precio_unitario,
        tipo=data.tipo,
    )
    db.add(servicio)
    db.flush()
    registrar_cambio_precio(
        db, tipo="servicio", referencia=servicio.nombre,
        precio_nuevo=data.precio_unitario, servicio_id=servicio.id,
        usuario=current_user, origen="manual",
    )
    db.commit()
    db.refresh(servicio)
    return servicio


@router.put("/{id}", response_model=schemas.ServicioOut)
def actualizar(id: str, data: schemas.ServicioUpdate, db: Session = Depends(get_db),
               current_user: models.Usuario = Depends(get_current_user)):
    _require_sdl_o_admin(current_user, db)
    servicio = db.query(models.Servicio).filter(models.Servicio.id == id).first()
    if not servicio:
        raise HTTPException(status_code=404, detail="Servicio no encontrado")
    payload = data.model_dump(exclude_none=True)
    precio_anterior = servicio.precio_unitario
    for field, value in payload.items():
        setattr(servicio, field, value)
    if "precio_unitario" in payload:
        registrar_cambio_precio(
            db, tipo="servicio", referencia=servicio.nombre,
            precio_nuevo=payload["precio_unitario"], precio_anterior=precio_anterior,
            servicio_id=servicio.id, usuario=current_user, origen="manual",
        )
    db.commit()
    db.refresh(servicio)
    return servicio


@router.delete("/{id}")
def eliminar(id: str, db: Session = Depends(get_db),
             current_user: models.Usuario = Depends(get_current_user)):
    _require_sdl_o_admin(current_user, db)
    servicio = db.query(models.Servicio).filter(models.Servicio.id == id).first()
    if not servicio:
        raise HTTPException(status_code=404, detail="Servicio no encontrado")
    servicio.activo = False
    db.commit()
    return {"detail": "Servicio desactivado"}


# ── Importación de servicios por Excel ───────────────────────
@router.get("/plantilla-importar")
def descargar_plantilla_servicios(_=Depends(require_admin)):
    """Excel vacío con los encabezados para importar servicios."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Servicios"
    headers = ["nombre", "tipo", "descripcion", "precio"]
    hf = Font(bold=True, color="FFFFFF")
    hfill = PatternFill(start_color="26326E", end_color="26326E", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")
    for i, name in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i, value=name)
        c.font = hf; c.fill = hfill; c.alignment = center
        ws.column_dimensions[c.column_letter].width = 26
    ws.cell(row=2, column=1, value="XGQ-20FII DE 20 KG")
    ws.cell(row=2, column=2, value="mantenimiento")
    ws.cell(row=2, column=3, value="El servicio de mantenimiento está diseñado para garantizar...")
    ws.cell(row=2, column=4, value=7150)
    nota = ("nombre = equipo/modelo · tipo = mantenimiento | puesta_en_marcha | otro · "
            "descripcion = qué incluye el servicio · precio en MXN. "
            "Un servicio se identifica por nombre + tipo (el mismo equipo puede tener "
            "mantenimiento y puesta en marcha).")
    ws.cell(row=3, column=1, value=nota)
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(headers))
    ws.cell(row=3, column=1).font = Font(italic=True, color="808080")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="plantilla_servicios.xlsx"'},
    )


@router.post("/importar")
def importar_servicios(
    file: UploadFile = File(...),
    confirmar: bool = False,   # False = vista previa; True = aplica
    db: Session = Depends(get_db),
    current_user: models.Usuario = Depends(require_admin),
):
    """Importa servicios desde Excel en dos fases (vista previa + confirmar).
    Identifica cada servicio por (nombre + tipo) y solo actualiza si hay cambios
    reales (precio o descripción)."""
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Solo se aceptan archivos .xlsx o .xls")
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file.file.read()), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="No se pudo leer el archivo Excel")
    rows = list(wb.active.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="El archivo está vacío")

    headers = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
    idx = {h: i for i, h in enumerate(headers) if h}
    faltan = {"nombre", "tipo", "precio"} - idx.keys()
    if faltan:
        raise HTTPException(
            status_code=400,
            detail=f"Formato incorrecto: faltan columnas ({', '.join(sorted(faltan))}). "
                   f"Descarga la plantilla de servicios. Encabezados: "
                   f"{', '.join(h for h in headers if h) or '(ninguno)'}",
        )

    def cel(row, k):
        i = idx.get(k)
        return row[i] if i is not None and i < len(row) else None

    def parse_precio(raw):
        if raw is None or raw == "":
            return None
        try:
            v = float(str(raw).replace(",", "").replace("$", "").strip())
            return v if v > 0 else None
        except (ValueError, TypeError):
            return None

    nuevos, actualizar, errores = [], [], []
    sin_cambios = 0
    for n, row in enumerate(rows[1:], start=2):
        nombre = str(cel(row, "nombre") or "").strip()
        tipo_raw = str(cel(row, "tipo") or "").strip().lower()
        desc = str(cel(row, "descripcion") or "").strip() or None
        precio = parse_precio(cel(row, "precio"))

        if not nombre and not tipo_raw and precio is None:
            continue  # fila en blanco
        if not nombre:
            errores.append(f"Fila {n}: falta el nombre"); continue
        tipo = _TIPO_ALIAS.get(tipo_raw)
        if tipo is None:
            errores.append(f"Fila {n} ({nombre}): tipo inválido '{tipo_raw}' "
                           f"(usa mantenimiento, puesta_en_marcha u otro)"); continue
        if precio is None:
            errores.append(f"Fila {n} ({nombre}): precio inválido o vacío"); continue

        ref = f"{nombre} [{tipo}]"
        existente = (db.query(models.Servicio)
                       .filter(func.lower(models.Servicio.nombre) == nombre.lower(),
                               models.Servicio.tipo == tipo).first())
        if existente:
            cambios = []
            if round(float(existente.precio_unitario), 2) != round(precio, 2):
                cambios.append(f"precio: {float(existente.precio_unitario):,.2f} → {precio:,.2f}")
            if desc is not None and (desc or None) != (existente.descripcion or None):
                cambios.append("descripción")
            if not existente.activo:
                cambios.append("reactivar")
            if cambios:
                actualizar.append({"servicio": existente, "desc": desc, "precio": precio,
                                   "cambios": cambios, "referencia": ref})
            else:
                sin_cambios += 1
        else:
            nuevos.append({"nombre": nombre, "tipo": tipo, "desc": desc,
                           "precio": precio, "referencia": ref})

    resumen = {"nuevos": len(nuevos), "actualizar": len(actualizar),
               "sin_cambios": sin_cambios, "errores": len(errores)}
    preview = {
        "resumen": resumen,
        "nuevos": [f"{x['referencia']} — {x['precio']:,.2f}" for x in nuevos[:100]],
        "actualizar": [{"item": x["referencia"], "cambios": x["cambios"]} for x in actualizar[:100]],
        "errores": errores[:100],
    }

    if not confirmar:
        return {**preview, "confirmado": False}

    try:
        for x in nuevos:
            s = models.Servicio(nombre=x["nombre"], tipo=x["tipo"],
                                descripcion=x["desc"], precio_unitario=x["precio"], activo=True)
            db.add(s); db.flush()
            registrar_cambio_precio(db, tipo="servicio", referencia=s.nombre,
                                    precio_nuevo=x["precio"], servicio_id=s.id,
                                    usuario=current_user, origen="importacion")
        for x in actualizar:
            s = x["servicio"]
            anterior = s.precio_unitario
            s.precio_unitario = x["precio"]
            s.activo = True
            if x["desc"] is not None:
                s.descripcion = x["desc"]
            db.flush()
            registrar_cambio_precio(db, tipo="servicio", referencia=s.nombre,
                                    precio_nuevo=x["precio"], precio_anterior=anterior,
                                    servicio_id=s.id, usuario=current_user, origen="importacion")
        db.commit()
    except Exception as ex:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"Error aplicando la importación: {ex}")

    return {**preview, "confirmado": True}
